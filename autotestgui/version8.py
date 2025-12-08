import tkinter as tk
from tkinter import ttk, filedialog, messagebox, simpledialog
import json
import threading
import time
import os
import re
from datetime import datetime
from test_step import TestStep

REPORT_OUTPUT_FOLDER = "TestReports"
os.makedirs(REPORT_OUTPUT_FOLDER, exist_ok=True)

combined_report_data = []
clipboard_step_data = []  # Now supports multiple steps

def save_combined_html():
    if not combined_report_data:
        return
    combined_html = """<html><head>
    <title>Test Summary</title>
    <style>
        body { font-family: Arial; padding: 20px; }
        h2 { color: #003366; }
        li { margin-bottom: 6px; }
    </style>
    </head><body><h1>Combined Test Summary</h1><hr>"""
    combined_html += "<hr>".join(combined_report_data)
    combined_html += "</body></html>"

    filename = os.path.join(REPORT_OUTPUT_FOLDER, "Combined_Test_Summary.html")
    with open(filename, "w", encoding="utf-8") as f:
        f.write(combined_html)

    messagebox.showinfo("Report Saved", f"Combined report saved at:\n{filename}")

class TestCaseFrame:
    def __init__(self, parent, name):
        global clipboard_step_data

        self.name = name
        self.frame = ttk.LabelFrame(parent, text=name, style="Case.TLabelframe")
        self.frame.pack(fill="x", padx=8, pady=6)

        self.steps = []
        self.selected_step_index = None

        # Action buttons at the top
        btns = ttk.Frame(self.frame, style="CaseInner.TFrame")
        btns.pack(pady=(10, 10), padx=12)

        ttk.Button(btns, text="‚ûï Add Step", command=self.add_step, style="Accent.TButton").grid(row=0, column=0, padx=2)
        ttk.Button(btns, text="üóë Clear", command=self.clear_steps, style="Ghost.TButton").grid(row=0, column=1, padx=2)
        ttk.Button(btns, text="‚ñ∂ Run", command=self.run, style="Accent.TButton").grid(row=0, column=2, padx=2)
        ttk.Button(btns, text="üìã Copy Checked", command=self.copy_checked_steps, style="Ghost.TButton").grid(row=0, column=3, padx=2)
        ttk.Button(btns, text="üóë Delete", command=self.delete_selected_step, style="Danger.TButton").grid(row=0, column=4, padx=2)
        ttk.Button(btns, text="üìã Paste", command=self.paste_step_from_clipboard, style="Ghost.TButton").grid(row=0, column=5, padx=2)
        ttk.Button(btns, text="‚úè Rename Step", command=self.rename_selected_step, style="Ghost.TButton").grid(row=0, column=6, padx=2)
        ttk.Button(btns, text="‚òë Select All", command=self.select_all_steps, style="Ghost.TButton").grid(row=0, column=7, padx=2)
        ttk.Button(btns, text="‚òê Unselect All", command=self.unselect_all_steps, style="Ghost.TButton").grid(row=0, column=8, padx=2)
        
        # Second row for category filtering
        btns2 = ttk.Frame(self.frame, style="CaseInner.TFrame")
        btns2.pack(pady=(0, 10), padx=12)
        
        ttk.Label(btns2, text="Filter by Category:", style="Case.TLabelframe.Label").grid(row=0, column=0, padx=2)
        ttk.Button(btns2, text="All", command=lambda: self.filter_by_category(None), style="Ghost.TButton").grid(row=0, column=1, padx=2)
        ttk.Button(btns2, text="Setup", command=lambda: self.filter_by_category("Setup"), style="Ghost.TButton").grid(row=0, column=2, padx=2)
        ttk.Button(btns2, text="Validation", command=lambda: self.filter_by_category("Validation"), style="Ghost.TButton").grid(row=0, column=3, padx=2)
        ttk.Button(btns2, text="Cleanup", command=lambda: self.filter_by_category("Cleanup"), style="Ghost.TButton").grid(row=0, column=4, padx=2)
        ttk.Button(btns2, text="Critical", command=lambda: self.filter_by_category("Critical"), style="Ghost.TButton").grid(row=0, column=5, padx=2)
        
        self.toggle_output_btn = ttk.Button(btns2, text="üìä Show Output", command=self.toggle_output, style="Ghost.TButton")
        self.toggle_output_btn.grid(row=0, column=6, padx=2)

        # Search bar
        self.search_var = tk.StringVar()
        search_frame = ttk.Frame(self.frame, style="CaseInner.TFrame")
        search_frame.pack(fill="x", padx=8, pady=(0, 6))
        search_entry = ttk.Entry(search_frame, textvariable=self.search_var, style="Search.TEntry")
        search_entry.pack(side="left", expand=True, fill="x")
        ttk.Button(search_frame, text="üîç Search", command=self.filter_steps, style="Ghost.TButton").pack(side="left", padx=(8, 0))
        ttk.Button(search_frame, text="‚úñ Clear", command=self.clear_filter, style="Ghost.TButton").pack(side="left", padx=6)

        # Scrollable steps container
        steps_container = ttk.Frame(self.frame, style="CaseInner.TFrame")
        steps_container.pack(fill="both", expand=True, padx=8)
        
        steps_canvas = tk.Canvas(steps_container, background="#ffffff", highlightthickness=0, bd=0, height=600)
        steps_scrollbar = ttk.Scrollbar(steps_container, orient="vertical", command=steps_canvas.yview)
        self.inner_frame = ttk.Frame(steps_canvas, style="CaseInner.TFrame")
        
        steps_canvas.pack(side="left", fill="both", expand=True)
        
        steps_canvas_window = steps_canvas.create_window((0, 0), window=self.inner_frame, anchor="nw", tags="steps_frame")
        
        self.inner_frame.bind(
            "<Configure>",
            lambda e: steps_canvas.configure(scrollregion=steps_canvas.bbox("all"))
        )
        
        steps_canvas.bind(
            "<Configure>",
            lambda e: steps_canvas.itemconfig("steps_frame", width=e.width)
        )
        
        # Mouse wheel scrolling for steps
        def _on_steps_mousewheel(event):
            steps_canvas.yview_scroll(int(-1 * (event.delta / 120)), "units")
        
        steps_canvas.bind_all("<MouseWheel>", _on_steps_mousewheel)
        steps_canvas.configure(yscrollcommand=steps_scrollbar.set)
        
        # Auto-hide scrollbar - show only on hover
        def _show_scrollbar(event=None):
            steps_scrollbar.pack(side="right", fill="y")
        
        def _hide_scrollbar(event=None):
            steps_scrollbar.pack_forget()
        
        steps_container.bind("<Enter>", _show_scrollbar)
        steps_container.bind("<Leave>", _hide_scrollbar)
        steps_canvas.bind("<Enter>", _show_scrollbar)
        
        # Start with scrollbar hidden
        _hide_scrollbar()

        # Output console inside scrollable container (hidden by default)
        self.output_frame = ttk.Frame(self.inner_frame, style="CaseInner.TFrame")
        self.output = tk.Text(self.output_frame, height=7, relief="solid", bg="#f8fafc", fg="#1e293b", wrap="word", borderwidth=1)
        self.output.configure(insertbackground="#1e293b", highlightthickness=0, bd=1, padx=8, pady=6, font=("Consolas", 9))
        self.output.pack(fill="both", expand=True)
        
        self.output_visible = False
        self.last_result = "Pending"

    def toggle_output(self):
        if self.output_visible:
            self.output_frame.pack_forget()
            self.toggle_output_btn.config(text="üìä Show Output")
            self.output_visible = False
        else:
            self.output_frame.pack(in_=self.inner_frame, fill="x", pady=(8, 0))
            self.toggle_output_btn.config(text="üìä Hide Output")
            self.output_visible = True

    def _bind_step_clicks(self, step, idx):
        handler = lambda e, i=idx: self.select_step(i)
        self._bind_widget_and_children(step.frame, handler)

    def _bind_widget_and_children(self, widget, handler):
        widget.bind("<Button-1>", handler)
        for child in widget.winfo_children():
            self._bind_widget_and_children(child, handler)

    def add_step(self):
        step = TestStep(self.inner_frame, len(self.steps)+1)
        step.frame.configure(style="Step.TLabelframe")
        step.frame.pack(fill="x", pady=8)
        self._bind_step_clicks(step, len(self.steps))
        
        # Bind category change to update colors
        step.category_dropdown.bind("<<ComboboxSelected>>", lambda e, s=step: self.update_step_category_color(s))
        
        self.steps.append(step)
    
    def update_step_category_color(self, step):
        """Update step frame border color based on category"""
        category_colors = {
            "General": "#64748b",
            "Setup": "#3b82f6", 
            "Validation": "#10b981",
            "Cleanup": "#f59e0b",
            "Critical": "#ef4444"
        }
        category = step.category.get()
        # We'll update the label color to indicate category
        color = category_colors.get(category, "#64748b")
        step.frame.configure(text=f"{step.step_name} [{category}]")

    def select_step(self, index):
        # Toggle selection: if clicking the same step, unselect it
        if self.selected_step_index == index:
            self.selected_step_index = None
            for step in self.steps:
                step.frame.config(style="Step.TLabelframe")
        else:
            self.selected_step_index = index
            for i, step in enumerate(self.steps):
                step.frame.config(style="StepSelected.TLabelframe" if i == index else "Step.TLabelframe")
    
    def copy_checked_steps(self):
        global clipboard_step_data
        clipboard_step_data = []
        step_names = []
        
        for idx, step in enumerate(self.steps):
            if step.is_checked.get():
                clipboard_step_data.append(step.get_step_data())
                step_name = step.get_step_data().get("name", f"Step {idx + 1}")
                step_names.append(f"  ‚Ä¢ {step_name}")
        
        if not clipboard_step_data:
            messagebox.showwarning("Copy Steps", "No steps are checked. Please check the steps you want to copy.")
            return
        
        # Create a detailed message showing which steps were copied
        if len(step_names) == 1:
            message = f"1 step copied to clipboard:\n\n{step_names[0]}"
        else:
            message = f"{len(step_names)} steps copied to clipboard:\n\n" + "\n".join(step_names)
        
        messagebox.showinfo("Steps Copied", message)
    
    def select_all_steps(self):
        """Check all step checkboxes"""
        for step in self.steps:
            step.is_checked.set(True)
    
    def unselect_all_steps(self):
        """Uncheck all step checkboxes"""
        for step in self.steps:
            step.is_checked.set(False)

    def paste_step_from_clipboard(self):
        global clipboard_step_data
        if not clipboard_step_data:
            messagebox.showwarning("Paste Step", "Clipboard is empty.")
            return

        for data in clipboard_step_data:
            step = TestStep(self.inner_frame, len(self.steps)+1)
            step.frame.configure(style="Step.TLabelframe")
            step.frame.pack(fill="x", pady=8)
            self._bind_step_clicks(step, len(self.steps))
            if "name" in data:
                step.step_name = data["name"]
                step.frame.config(text=data["name"])
            step.step_type.set(data["type"])
            step.show_fields()

            for key, widget in step.details.items():
                value = data["details"].get(key)
                if value is not None:
                    if isinstance(widget, ttk.Combobox):
                        widget.set(value)
                    elif isinstance(widget, tk.Entry):
                        widget.delete(0, tk.END)
                        widget.insert(0, value)
                    elif key == "from_files":
                        step.details["from_files"] = value
                        if "from" in step.details:
                            step.details["from"].delete(0, tk.END)
                            step.details["from"].insert(0, ", ".join(value))

            if data["type"] == "Check Database Entry":
                for idx, col in enumerate(data["details"].get("columns", [])):
                    if idx >= len(step.column_entries):
                        step.add_column_entry()
                    name_entry, op_entry, val_entry = step.column_entries[idx]
                    name_entry.insert(0, col.get("column", ""))
                    op_entry.set(col.get("operator", "="))
                    val_entry.insert(0, col.get("value", ""))

            self.steps.append(step)

    def delete_selected_step(self):
        if self.selected_step_index is None or self.selected_step_index >= len(self.steps):
            messagebox.showwarning("Delete Step", "No step selected.")
            return
        step = self.steps.pop(self.selected_step_index)
        step.frame.destroy()
        self.selected_step_index = None
        # Rebind remaining steps
        for idx, step in enumerate(self.steps):
            self._bind_step_clicks(step, idx)
        self.select_step(None)

    def clear_steps(self):
        for step in self.steps:
            step.frame.destroy()
        self.steps.clear()
        self.selected_step_index = None
        self.output.delete("1.0", tk.END)

    def rename_selected_step(self):
        if self.selected_step_index is None or self.selected_step_index >= len(self.steps):
            messagebox.showwarning("Rename Step", "No step selected.")
            return
        step = self.steps[self.selected_step_index]
        new_name = simpledialog.askstring("Rename Step", f"Enter new name for Step {self.selected_step_index + 1}:", initialvalue=step.step_name)
        if new_name:
            step.step_name = new_name
            category = step.category.get()
            step.frame.config(text=f"{new_name} [{category}]")

    def rename_test_case(self):
        new_name = simpledialog.askstring("Rename Test Case", "Enter new test case name:", initialvalue=self.name)
        if new_name:
            self.name = new_name
            self.frame.config(text=new_name)

    def get_data(self):
        return [step.get_step_data() for step in self.steps]

    def load_data(self, steps_data):
        self.clear_steps()
        for step_data in steps_data:
            step = TestStep(self.inner_frame, len(self.steps)+1)
            step.frame.configure(style="Step.TLabelframe")
            step.frame.pack(fill="x", pady=8)
            self._bind_step_clicks(step, len(self.steps))
            if "name" in step_data:
                step.step_name = step_data["name"]
                step.frame.config(text=step_data["name"])
            step.step_type.set(step_data["type"])
            step.show_fields()
            for key, widget in step.details.items():
                value = step_data["details"].get(key)
                if value is not None:
                    if isinstance(widget, ttk.Combobox):
                        widget.set(value)
                    elif isinstance(widget, tk.Entry):
                        widget.delete(0, tk.END)
                        widget.insert(0, value)
                    elif key == "from_files":
                        step.details["from_files"] = value
                        if "from" in step.details:
                            step.details["from"].delete(0, tk.END)
                            step.details["from"].insert(0, ", ".join(value))

            if step_data["type"] == "Check Database Entry":
                for idx, col in enumerate(step_data["details"].get("columns", [])):
                    if idx >= len(step.column_entries):
                        step.add_column_entry()
                    name_entry, op_entry, val_entry = step.column_entries[idx]
                    name_entry.insert(0, col.get("column", ""))
                    op_entry.set(col.get("operator", "="))
                    val_entry.insert(0, col.get("value", ""))

            self.steps.append(step)

    def filter_steps(self):
        search = self.search_var.get().strip().lower()
        for step in self.steps:
            match = search in step.step_type.get().lower()
            step.frame.pack_forget()
            if match or not search:
                step.frame.pack(fill="x", pady=8)

    def clear_filter(self):
        self.search_var.set("")
        self.filter_steps()
    
    def filter_by_category(self, category=None):
        """Filter steps by category. If category is None, show all steps."""
        for step in self.steps:
            step.frame.pack_forget()
            if category is None or step.category.get() == category:
                step.frame.pack(fill="x", pady=8)


    def run(self):
        def execute():
            import shutil
            try:
                import pyodbc  # type: ignore
            except ImportError:
                self.output.insert(tk.END, "‚ùå pyodbc is not installed. Install it and retry.\n")
                self.last_result = "FAIL"
                return
            from test_step import load_db_config
            from datetime import datetime, timedelta
            self.output.delete("1.0", tk.END)
            self.output.insert(tk.END, f"‚ñ∂ Running test case: {self.name}\n")
            success = True
            timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

            log_lines = [f"[{timestamp}] Running {self.name}"]
            html_report = [f"<h2>Test Case: {self.name}</h2><ul>"]
            
            # Track execution metrics
            total_steps = len(self.steps)
            executed_steps = 0
            skipped_steps = 0
            previous_step_passed = None

            for i, step in enumerate(self.steps, 1):
                try:
                    step_data = step.get_step_data()
                    run_condition = step_data.get("run_condition", "Always")
                    category = step_data.get("category", "General")
                    
                    # Check conditional execution
                    should_run = True
                    skip_reason = ""
                    
                    if run_condition == "Skip":
                        should_run = False
                        skip_reason = "Marked as Skip"
                    elif run_condition == "If Previous Passed" and previous_step_passed is False:
                        should_run = False
                        skip_reason = "Previous step failed"
                    elif run_condition == "If Previous Failed" and previous_step_passed is True:
                        should_run = False
                        skip_reason = "Previous step passed"
                    elif run_condition == "If Previous Passed" and previous_step_passed is None and i > 1:
                        should_run = False
                        skip_reason = "Previous step had no result"
                    
                    if not should_run:
                        step_name = step_data.get('name', f'Step {i}')
                        msg = f"‚è≠ Step {i}: {step_name} [{category}]: SKIPPED - {skip_reason}"
                        self.output.insert(tk.END, msg + "\n")
                        log_lines.append(f"[{datetime.now()}] {msg}")
                        html_report.append(f"<li style='color: #888;'>{msg}</li>")
                        skipped_steps += 1
                        continue
                    
                    executed_steps += 1
                    step_start_time = time.time()
                    
                    delay = int(step_data["details"].get("step_delay", 0))
                    if delay > 0:
                        msg = f"‚è± Waiting {delay} seconds before Step {i}"
                        self.output.insert(tk.END, msg + "\n")
                        self.output.update()
                        time.sleep(delay)
                        log_lines.append(f"[{datetime.now()}] {msg}")

                    step_name = step_data.get('name', f'Step {i}')
                    msg = f"‚û° Step {i}: {step_name} [{category}]: {step_data['type']}"
                    self.output.insert(tk.END, msg + "\n")
                    details = step_data["details"]
                    log_lines.append(f"[{datetime.now()}] {msg}")
                    passed = True

                    if step_data["type"] == "Copy File":
                        from_files = details.get("from_files", [])
                        dest = details.get("to", "")
                        
                        if not from_files:
                            msg = f"‚ùå No source files specified"
                            passed = False
                            self.output.insert(tk.END, msg + "\n")
                            log_lines.append(f"[{datetime.now()}] {msg}")
                        elif not dest:
                            msg = f"‚ùå No destination path specified"
                            passed = False
                            self.output.insert(tk.END, msg + "\n")
                            log_lines.append(f"[{datetime.now()}] {msg}")
                        else:
                            for src in from_files:
                                if os.path.isfile(src):
                                    shutil.copy(src, dest)
                                    msg = f"‚úÖ Copied '{src}' to '{dest}'"
                                else:
                                    msg = f"‚ùå Source file not found: {src}"
                                    passed = False
                                self.output.insert(tk.END, msg + "\n")
                                log_lines.append(f"[{datetime.now()}] {msg}")



                    elif step_data["type"] == "Check Log File":
                        path = details.get("log_file_path")
                        log_type = details.get("log_type", "").lower()
                        search = details.get("search", "").strip().lower()
                        search = search.encode('unicode_escape').decode().lower()
                        log_delay = int(details.get("delay", 0))
                        duration = int(details.get("duration", 0))  # in minutes
                        custom_format = details.get("timestamp_format")  # optional override

                        timestamp_patterns = [
                            (r"^\d{4}-\d{2}-\d{2} \d{2}:\d{2}:\d{2},\d{3}", "%Y-%m-%d %H:%M:%S,%f"),  # <-- Add this
                            (r"^\d{2}\.\d{2}\.\d{4} \d{2}:\d{2}:\d{2}\.\d{3}", "%d.%m.%Y %H:%M:%S.%f"),
                            (r"^\d{4}-\d{2}-\d{2} \d{2}:\d{2}:\d{2}\.\d{3}", "%Y-%m-%d %H:%M:%S.%f"),
                            (r"^\d{4}-\d{2}-\d{2} \d{2}:\d{2}:\d{2}\.\d{3}(?=\s|$)", "%Y-%m-%d %H:%M:%S.%f"),
                            (r"^\d{4}-\d{2}-\d{2} \d{2}:\d{2}:\d{2}", "%Y-%m-%d %H:%M:%S"),
                            (r"^\d{2}-\d{2}-\d{4} \d{2}:\d{2}:\d{2}", "%d-%m-%Y %H:%M:%S"),
                            (r"^\d{2}/\d{2}/\d{4} \d{2}:\d{2}:\d{2}", "%d/%m/%Y %H:%M:%S"),
                            (r"^\d{4}/\d{2}/\d{2} \d{2}:\d{2}:\d{2}", "%Y/%m/%d %H:%M:%S"),
                            (r"^\d{4}-\d{2}-\d{2} \d{2}:\d{2}", "%Y-%m-%d %H:%M"),
                        ]

                        if custom_format:
                            timestamp_patterns.insert(0, (r"^.*", custom_format))  # allow full line parsing

                        if log_delay > 0:
                            msg = f"‚è≥ Waiting {log_delay}s before log check"
                            self.output.insert(tk.END, msg + "\n")
                            self.output.update()
                            time.sleep(log_delay)
                            log_lines.append(f"[{datetime.now()}] {msg}")

                        if not os.path.exists(path):
                            msg = f"‚ùå Log file not found: {path}"
                            self.output.insert(tk.END, msg + "\n")
                            log_lines.append(f"[{datetime.now()}] {msg}")
                            passed = False
                        else:
                            now = datetime.now()
                            time_threshold = now - timedelta(minutes=duration)
                            matched = []

                            with open(path, "r", encoding="utf-8", errors="ignore") as f:
                                lines = f.readlines()

                            for line in lines:
                                log_time = None
                                for pattern, fmt in timestamp_patterns:
                                    match = re.match(pattern, line)
                                    if match:
                                        timestamp_str = match.group()
                                        try:
                                            if "%f" in fmt and "," in timestamp_str:
                                                timestamp_str = timestamp_str.replace(",", ".")
                                            log_time = datetime.strptime(timestamp_str, fmt)
                                            break
                                        except ValueError:
                                            continue

                                if not log_time:
                                    continue

                                # Time-based filter
                                if duration > 0 and log_time < time_threshold:
                                    continue

                                # Case-insensitive match checks
                                normalized_line = line.lower()

                                # Log type filter
                                if log_type and not re.search(rf"\b{re.escape(log_type.lower())}\b", normalized_line):
                                    continue

                                # Normalize inputs to avoid escape issues
                                # Normalize slashes
                                normalized_line = line.strip().lower().replace("\\", "/").replace("//", "/")
                                normalized_search = search.strip().lower().replace("\\", "/").replace("//", "/")

                                print("Searching for:", repr(normalized_search))
                                print("In log line:", repr(normalized_line))

                                if normalized_search and normalized_search not in normalized_line:
                                    continue

                                matched.append(line)

                                # if duration == 0 or log_time >= time_threshold:
                                #     if (log_type in line.lower()) and (search in line.lower() if search else True):
                                #         matched.append(line)

                            if matched:
                                msg = f"‚úÖ Found {len(matched)} matching log lines (last {duration} min). Last 3:"
                                passed = True
                            else:
                                msg = "‚ùå No entries found that match the search parameters.."
                                passed = False

                            self.output.insert(tk.END, msg + "\n")
                            log_lines.append(f"[{datetime.now()}] {msg}")
                            for line in matched[-3:]:
                                self.output.insert(tk.END, f"   ‚û§ {line.strip()}\n")

                    elif step_data["type"] == "Check Database Entry":
                        config = load_db_config()
                        if not config:
                            msg = "‚ùå DB config not loaded"
                            passed = False
                        else:
                            conn = pyodbc.connect(
                                f"DRIVER={{SQL Server}};SERVER={config['server']};DATABASE={config['database']};"
                                f"UID={config['username']};PWD={config['password']}"
                            )
                            cursor = conn.cursor()
                            where = " AND ".join(
                                f"{entry['column']} {entry['operator']} '{entry['value']}'" for entry in
                                details.get("columns", []))
                            sql = f"SELECT COUNT(*) FROM {details['table']} WHERE {where}"
                            msg = f"üßæ Executing SQL: {sql}"
                            self.output.insert(tk.END, msg + "\n")
                            log_lines.append(f"[{datetime.now()}] {msg}")
                            cursor.execute(sql)
                            count = cursor.fetchone()[0]
                            if count > 0:
                                msg = f"‚úÖ Found {count} matching rows."
                            else:
                                msg = "‚ùå No matching records."
                                passed = False
                            self.output.insert(tk.END, msg + "\n")
                            log_lines.append(f"[{datetime.now()}] {msg}")
                            cursor.close()
                            conn.close()

                    # Calculate execution time for this step
                    step_end_time = time.time()
                    step_execution_time = step_end_time - step_start_time
                    step.execution_time = step_execution_time
                    step.last_result = "PASS" if passed else "FAIL"
                    previous_step_passed = passed
                    
                    result_msg = f"{'‚úîÔ∏è' if passed else '‚ùå'} Step {i}: {step_name} [{category}] {'passed' if passed else 'failed'} ({step_execution_time:.2f}s)"
                    self.output.insert(tk.END, result_msg + "\n")
                    
                    # Enhanced HTML reporting with color coding
                    color = "#10b981" if passed else "#ef4444"
                    html_report.append(f"<li style='color: {color};'>{result_msg}</li>")
                    log_lines.append(f"[{datetime.now()}] {result_msg}")
                    self.output.update()
                    if not passed:
                        success = False

                except Exception as e:
                    step_end_time = time.time()
                    step_execution_time = step_end_time - step_start_time if 'step_start_time' in locals() else 0
                    step.execution_time = step_execution_time
                    step.last_result = "ERROR"
                    previous_step_passed = False
                    executed_steps += 1
                    step_name = step.get_step_data().get('name', f'Step {i}')
                    
                    msg = f"‚ùå Error in Step {i}: {step_name}: {e} ({step_execution_time:.2f}s)"
                    self.output.insert(tk.END, msg + "\n")
                    html_report.append(f"<li style='color: #ef4444;'>{msg}</li>")
                    log_lines.append(f"[{datetime.now()}] {msg}")
                    success = False

            # Calculate total execution time
            total_execution_time = sum(step.execution_time for step in self.steps if hasattr(step, 'execution_time'))
            
            # Generate execution summary
            final_msg = f"\n{'‚úÖ PASSED' if success else '‚ùå FAILED'}"
            summary = f"\nüìä Execution Summary:\n"
            summary += f"   ‚Ä¢ Total Steps: {total_steps}\n"
            summary += f"   ‚Ä¢ Executed: {executed_steps}\n"
            summary += f"   ‚Ä¢ Skipped: {skipped_steps}\n"
            summary += f"   ‚Ä¢ Total Time: {total_execution_time:.2f}s\n"
            
            self.output.insert(tk.END, final_msg + "\n")
            self.output.insert(tk.END, summary)
            
            # Enhanced HTML summary
            html_report.append(f"</ul>")
            html_report.append(f"<div style='margin-top: 20px; padding: 15px; background: #f8fafc; border-radius: 8px;'>")
            html_report.append(f"<h3 style='margin-top: 0;'>üìä Execution Summary</h3>")
            html_report.append(f"<p><strong>Status:</strong> <span style='color: {'#10b981' if success else '#ef4444'};'>{('PASSED' if success else 'FAILED')}</span></p>")
            html_report.append(f"<p><strong>Total Steps:</strong> {total_steps}</p>")
            html_report.append(f"<p><strong>Executed:</strong> {executed_steps}</p>")
            html_report.append(f"<p><strong>Skipped:</strong> {skipped_steps}</p>")
            html_report.append(f"<p><strong>Total Execution Time:</strong> {total_execution_time:.2f}s</p>")
            html_report.append(f"<p><strong>Timestamp:</strong> {timestamp}</p>")
            html_report.append(f"</div>")
            
            log_lines.append(f"[{datetime.now()}] {final_msg}")
            log_lines.append(f"[{datetime.now()}] {summary}")
            self.last_result = "PASS" if success else "FAIL"

            safe_name = re.sub(r'[^a-zA-Z0-9_\-]', '_', self.name)
            log_path = os.path.join(REPORT_OUTPUT_FOLDER, f"log_{safe_name}.txt")
            html_path = os.path.join(REPORT_OUTPUT_FOLDER, f"report_{safe_name}.html")

            with open(log_path, "w", encoding="utf-8") as log_file:
                log_file.write("\n".join(log_lines))

            try:
                with open(html_path, "w", encoding="utf-8") as html_file:
                    html_file.write("<html><body>" + "\n".join(html_report) + "</body></html>")
                    combined_report_data.append("".join(html_report))

            except Exception as e:
                self.output.insert(tk.END, f"‚ùå Failed to write HTML report: {e}\n")

            save_combined_html()

        threading.Thread(target=execute, daemon=True).start()


class TestCaseGUI:
    def __init__(self, root):
        self.root = root
        self.root.title("Multi Test Case Runner")

        self.colors = {
            "bg": "#f0f4f8",
            "card": "#ffffff",
            "surface": "#f8fafc",
            "accent": "#3b82f6",
            "accent_hover": "#2563eb",
            "accent_pressed": "#1d4ed8",
            "success": "#10b981",
            "success_hover": "#059669",
            "danger": "#ef4444",
            "danger_hover": "#dc2626",
            "muted": "#64748b",
            "heading": "#1e293b",
            "border": "#e2e8f0",
            "shadow": "#cbd5e1",
        }

        style = ttk.Style(root)
        style.theme_use("clam")

        root.configure(bg=self.colors["bg"])
        root.option_add("*Font", "Arial 10")

        style.configure("App.TFrame", background=self.colors["bg"])
        style.configure("Toolbar.TFrame", background=self.colors["bg"], padding=6)
        style.configure("Main.TFrame", background=self.colors["bg"])
        style.configure(
            "Case.TLabelframe",
            background=self.colors["card"],
            borderwidth=1,
            bordercolor=self.colors["border"],
            padding=10,
            relief="flat",
        )
        style.configure(
            "Case.TLabelframe.Label",
            background=self.colors["card"],
            foreground=self.colors["heading"],
            font=("Segoe UI Semibold", 12),
        )
        style.configure("CaseInner.TFrame", background=self.colors["card"])
        style.configure("StepInner.TFrame", background=self.colors["surface"])

        style.configure(
            "Step.TLabelframe",
            background=self.colors["surface"],
            borderwidth=1,
            bordercolor=self.colors["border"],
            padding=10,
            relief="flat",
        )
        style.configure(
            "Step.TLabelframe.Label",
            background=self.colors["surface"],
            foreground=self.colors["heading"],
        )
        style.configure(
            "StepSelected.TLabelframe",
            background="#dbeafe",
            borderwidth=2,
            bordercolor=self.colors["accent"],
            padding=10,
            relief="flat",
        )
        style.configure(
            "StepSelected.TLabelframe.Label",
            background="#dbeafe",
            foreground=self.colors["accent"],
            font=("Segoe UI Semibold", 10),
        )
        style.configure("Step.TLabel", background=self.colors["surface"], foreground=self.colors["heading"])

        style.configure(
            "Accent.TButton",
            background=self.colors["accent"],
            foreground="white",
            padding=(10, 6),
            borderwidth=0,
            relief="flat",
            font=("Segoe UI Semibold", 9),
        )
        style.map(
            "Accent.TButton",
            background=[("active", self.colors["accent_hover"]), ("pressed", self.colors["accent_pressed"])],
            relief=[("pressed", "sunken")],
        )

        style.configure(
            "Ghost.TButton",
            background=self.colors["card"],
            foreground=self.colors["accent"],
            padding=(10, 6),
            borderwidth=1,
            relief="flat",
            font=("Segoe UI Semibold", 9),
        )
        style.map(
            "Ghost.TButton",
            background=[("active", self.colors["surface"]), ("pressed", "#e2e8f0")],
            bordercolor=[("active", self.colors["accent"]), ("!active", self.colors["border"])],
        )

        style.configure(
            "Danger.TButton",
            background=self.colors["danger"],
            foreground="white",
            padding=(10, 6),
            borderwidth=0,
            relief="flat",
            font=("Segoe UI Semibold", 9),
        )
        style.map(
            "Danger.TButton",
            background=[("active", self.colors["danger_hover"]), ("pressed", "#b91c1c")],
            relief=[("pressed", "sunken")],
        )

        style.configure(
            "Success.TButton",
            background=self.colors["success"],
            foreground="white",
            padding=(10, 6),
            borderwidth=0,
            relief="flat",
            font=("Segoe UI Semibold", 9),
        )
        style.map(
            "Success.TButton",
            background=[("active", self.colors["success_hover"]), ("pressed", "#047857")],
            relief=[("pressed", "sunken")],
        )

        style.configure("TCombobox", padding=6, fieldbackground=self.colors["card"], background=self.colors["card"], foreground=self.colors["heading"])
        style.map(
            "TCombobox",
            fieldbackground=[("readonly", self.colors["card"])],
            selectbackground=[("readonly", self.colors["surface"])],
            selectforeground=[("readonly", self.colors["heading"])],
            foreground=[("active", self.colors["heading"])],
        )

        style.configure(
            "Step.TCombobox",
            padding=6,
            fieldbackground=self.colors["surface"],
            background=self.colors["surface"],
            foreground=self.colors["heading"],
        )
        style.map(
            "Step.TCombobox",
            fieldbackground=[("readonly", self.colors["surface"])],
            selectbackground=[("readonly", "#e0f2fe")],
            selectforeground=[("readonly", self.colors["heading"])],
            foreground=[("active", self.colors["heading"])],
        )

        style.configure("TEntry", padding=8, fieldbackground=self.colors["surface"], background=self.colors["surface"], relief="flat", borderwidth=1)
        style.configure("Search.TEntry", padding=8, fieldbackground=self.colors["card"], background=self.colors["card"], foreground=self.colors["muted"], relief="solid", borderwidth=1)
        style.configure("TLabel", background=self.colors["card"], foreground=self.colors["heading"])
        style.configure("Header.TLabel", background=self.colors["card"], foreground=self.colors["heading"])
        style.configure("Header.TFrame", background=self.colors["card"], padding=8)

        # Modern scrollbar styling
        style.configure("Vertical.TScrollbar", 
            background=self.colors["border"],
            troughcolor=self.colors["surface"],
            borderwidth=0,
            arrowsize=0,
            width=8)
        style.map("Vertical.TScrollbar",
            background=[("active", self.colors["accent"]), ("!active", self.colors["muted"])])

        self.case_frames = {}

        self.dropdown_var = tk.StringVar()

        header_shadow = ttk.Frame(root, style="App.TFrame", height=2)
        header_shadow.pack(fill="x")
        
        header = ttk.Frame(root, style="Header.TFrame")
        header.pack(fill="x", padx=0, pady=0)
        header_label = ttk.Label(header, text="üß™... ", style="Header.TLabel")
        header_label.configure(font=("Arial", 13, "bold"))
        header_label.pack(side="left", padx=4)
        
        # Action buttons in header
        btns = ttk.Frame(header, style="Header.TFrame")
        btns.pack(side="left", padx=10)

        ttk.Button(btns, text="‚ûï New Case", command=self.new_case, style="Accent.TButton").grid(row=0, column=0, padx=2)
        ttk.Button(btns, text="‚úèÔ∏è Rename", command=self.rename_case, style="Ghost.TButton").grid(row=0, column=1, padx=2)
        ttk.Button(btns, text="üóëÔ∏è Delete", command=self.delete_case, style="Danger.TButton").grid(row=0, column=2, padx=2)
        ttk.Button(btns, text="üì§ Export All", command=self.export_all, style="Ghost.TButton").grid(row=0, column=3, padx=2)
        ttk.Button(btns, text="üì• Import All", command=self.import_all, style="Ghost.TButton").grid(row=0, column=4, padx=2)
        ttk.Button(btns, text="üìä Export Reports to Excel", command=self.export_reports_to_excel, style="Ghost.TButton").grid(row=0, column=5, padx=2)
        ttk.Button(btns, text="‚ñ∂Ô∏è Run All Sequential", command=self.run_all_cases, style="Accent.TButton").grid(row=0, column=6, padx=2)
        ttk.Button(btns, text="‚ö° Run All Parallel", command=self.run_all_cases_parallel, style="Accent.TButton").grid(row=0, column=7, padx=2)
        
        dropdown_frame = ttk.Frame(header, style="Header.TFrame")
        dropdown_frame.pack(side="right", padx=4)
        ttk.Label(dropdown_frame, text="Active Test Case:", background=self.colors["card"], foreground=self.colors["muted"], font=("Segoe UI", 9)).pack(side="left", padx=(0, 6))
        self.dropdown = ttk.Combobox(dropdown_frame, textvariable=self.dropdown_var, state="readonly", width=22, font=("Segoe UI", 9))
        self.dropdown.bind("<<ComboboxSelected>>", self.switch_case)
        self.dropdown.pack(side="right")

        separator = ttk.Frame(root, height=1, style="App.TFrame")
        separator.pack(fill="x")
        separator.configure(relief="sunken", borderwidth=1)

        root.minsize(1080, 760)

        self.main_area = ttk.Frame(root, style="App.TFrame")
        self.main_area.pack(fill="both", expand=True)

        # Direct container without outer scrolling (scrolling is inside each test case)
        self.case_container = ttk.Frame(self.main_area, style="App.TFrame")
        self.case_container.pack(fill="both", expand=True)

    def new_case(self):
        name = f"TestCase{len(self.case_frames) + 1}"
        frame = TestCaseFrame(self.case_container, name)
        self.case_frames[name] = frame
        self.update_dropdown()
        self.dropdown.set(name)
        self.switch_case()

    def update_dropdown(self):
        self.dropdown["values"] = list(self.case_frames.keys())

    def switch_case(self, event=None):
        for widget in self.case_container.winfo_children():
            widget.pack_forget()
        name = self.dropdown_var.get()
        if name in self.case_frames:
            self.case_frames[name].frame.pack(fill="both", expand=True)

    def rename_case(self):
        old_name = self.dropdown_var.get()
        if old_name:
            new_name = simpledialog.askstring("Rename Test Case", "Enter new name:", initialvalue=old_name)
            if new_name and new_name not in self.case_frames:
                self.case_frames[new_name] = self.case_frames.pop(old_name)
                self.case_frames[new_name].name = new_name
                self.case_frames[new_name].frame.config(text=new_name)
                self.update_dropdown()
                self.dropdown.set(new_name)
                self.switch_case()

    def delete_case(self):
        name = self.dropdown_var.get()
        if name and messagebox.askyesno("Delete Test Case", f"Delete {name}?"):
            self.case_frames[name].frame.destroy()
            del self.case_frames[name]
            self.update_dropdown()
            if self.case_frames:
                first = list(self.case_frames.keys())[0]
                self.dropdown.set(first)
                self.switch_case()
            else:
                self.dropdown.set("")

    def export_all(self):
        data = {name: frame.get_data() for name, frame in self.case_frames.items()}
        file = filedialog.asksaveasfilename(defaultextension=".json")
        if file:
            with open(file, "w") as f:
                json.dump(data, f, indent=2)
            messagebox.showinfo("Exported", f"Exported all test cases to {file}")

    def import_all(self):
        file = filedialog.askopenfilename(filetypes=[("JSON Files", "*.json")])
        if not file:
            return
        try:
            with open(file) as f:
                data = json.load(f)
            self.case_frames.clear()
            for widget in self.case_container.winfo_children():
                widget.destroy()

            for name, steps in data.items():
                frame = TestCaseFrame(self.case_container, name)
                frame.load_data(steps)
                self.case_frames[name] = frame

            self.update_dropdown()
            names = list(self.case_frames.keys())
            if names:
                self.dropdown.set(names[0])
                self.switch_case()
        except Exception as e:
            messagebox.showerror("Import Failed", str(e))

    def run_all_cases(self):
        def run_all():
            start_time = time.time()
            results = []
            for name, frame in self.case_frames.items():
                frame.run()
                while frame.last_result == "Pending":
                    time.sleep(0.2)
                results.append(f"{name}: {frame.last_result}")
            end_time = time.time()
            total_time = end_time - start_time
            
            summary = "\n".join(results)
            summary += f"\n\nTotal execution time: {total_time:.2f}s"
            with open("test_summary.txt", "w") as f:
                f.write(summary)
            messagebox.showinfo("Summary Report", f"‚úÖ Completed test cases (Sequential):\n{summary}\nSaved to test_summary.txt")

        threading.Thread(target=run_all, daemon=True).start()
    
    def run_all_cases_parallel(self):
        """Run all test cases in parallel for faster execution"""
        def run_parallel():
            import concurrent.futures
            start_time = time.time()
            results = []
            
            def run_single_case(name, frame):
                frame.run()
                while frame.last_result == "Pending":
                    time.sleep(0.2)
                return f"{name}: {frame.last_result}"
            
            with concurrent.futures.ThreadPoolExecutor(max_workers=len(self.case_frames)) as executor:
                futures = {executor.submit(run_single_case, name, frame): name 
                          for name, frame in self.case_frames.items()}
                
                for future in concurrent.futures.as_completed(futures):
                    try:
                        result = future.result()
                        results.append(result)
                    except Exception as e:
                        case_name = futures[future]
                        results.append(f"{case_name}: ERROR - {e}")
            
            end_time = time.time()
            total_time = end_time - start_time
            
            summary = "\n".join(results)
            summary += f"\n\nTotal execution time (Parallel): {total_time:.2f}s"
            with open("test_summary_parallel.txt", "w") as f:
                f.write(summary)
            messagebox.showinfo("Summary Report", f"‚úÖ Completed test cases (Parallel):\n{summary}\nSaved to test_summary_parallel.txt")
        
        threading.Thread(target=run_parallel, daemon=True).start()
    
    def export_reports_to_excel(self):
        """Export test execution reports to Excel"""
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment
        except ImportError:
            messagebox.showerror("Missing Library", "openpyxl is not installed.\nInstall it with: pip install openpyxl")
            return
        
        if not self.case_frames:
            messagebox.showwarning("No Data", "No test cases to export.")
            return
        
        file = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel Files", "*.xlsx")],
            initialfile=f"Test_Report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        )
        
        if not file:
            return
        
        wb = openpyxl.Workbook()
        wb.remove(wb.active)  # Remove default sheet
        
        # Create summary sheet
        summary_sheet = wb.create_sheet("Summary")
        summary_sheet.append(["Test Case", "Total Steps", "Status", "Execution Time (s)"])
        
        # Style headers
        for cell in summary_sheet[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="3B82F6", end_color="3B82F6", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")
        
        # Add data for each test case
        for name, frame in self.case_frames.items():
            total_steps = len(frame.steps)
            status = getattr(frame, 'last_result', 'Not Run')
            total_time = sum(getattr(step, 'execution_time', 0) for step in frame.steps)
            
            summary_sheet.append([name, total_steps, status, f"{total_time:.2f}"])
            
            # Create detailed sheet for each test case
            sheet = wb.create_sheet(name[:31])  # Excel sheet name limit
            sheet.append(["Step #", "Name", "Type", "Category", "Condition", "Status", "Time (s)"])
            
            # Style headers
            for cell in sheet[1]:
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="10B981", end_color="10B981", fill_type="solid")
                cell.alignment = Alignment(horizontal="center")
            
            # Add step details
            for i, step in enumerate(frame.steps, 1):
                step_data = step.get_step_data()
                status = getattr(step, 'last_result', 'Not Run')
                exec_time = getattr(step, 'execution_time', 0)
                
                row = [
                    i,
                    step_data.get('name', f'Step {i}'),
                    step_data.get('type', 'N/A'),
                    step_data.get('category', 'General'),
                    step_data.get('run_condition', 'Always'),
                    status,
                    f"{exec_time:.2f}"
                ]
                sheet.append(row)
                
                # Color code status
                status_cell = sheet.cell(row=i+1, column=6)
                if status == "PASS":
                    status_cell.fill = PatternFill(start_color="D1FAE5", end_color="D1FAE5", fill_type="solid")
                elif status == "FAIL" or status == "ERROR":
                    status_cell.fill = PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid")
        
        # Auto-adjust column widths
        for sheet in wb.worksheets:
            for column in sheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(cell.value)
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                sheet.column_dimensions[column_letter].width = adjusted_width
        
        wb.save(file)
        messagebox.showinfo("Export Successful", f"Reports exported to:\n{file}")


def launch_app():
    root = tk.Tk()
    app = TestCaseGUI(root)
    root.geometry("950x750")
    root.title("Vasu Test Case Builder and Runner")
    root.state('zoomed')
    root.mainloop()


if __name__ == "__main__":
    launch_app()